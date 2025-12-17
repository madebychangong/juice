"""
재고표 엑셀 생성 스크립트
본두리편의점 재고실사.xlsx 파일 내에 재고표_출력 시트를 자동 생성
재고실사양식표 콜라 시트 형식과 동일하게 (열 너비, 행 높이 포함)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime


def format_date(date_value):
    """소비기한을 월/일 형식으로 변환"""
    if pd.isna(date_value):
        return ""

    if isinstance(date_value, str):
        date_obj = pd.to_datetime(date_value)
    else:
        date_obj = date_value

    return date_obj.strftime('%m/%d')


def format_quantity_date(quantity, date_value):
    """수량(월/일) 형식으로 변환"""
    if pd.isna(quantity) or quantity == '' or quantity == 0:
        return ""

    date_str = format_date(date_value)
    if date_str == "":
        return ""

    return f"{int(quantity)}({date_str})"


def copy_column_widths(source_sheet, target_sheet):
    """열 너비 복사"""
    for col in ['A', 'B', 'C', 'D', 'E']:
        if col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width


def create_inventory_sheet_in_excel(input_file, template_file):
    """
    엑셀 파일 내에 재고표 시트 생성

    Args:
        input_file: 본두리편의점 재고실사.xlsx
        template_file: 재고실사양식표.xlsx
    """
    print("📋 재고표 시트 생성 시작...")
    print(f"   입력: {input_file}")
    print(f"   양식: {template_file}")

    # 1. 정리표 데이터 읽기
    df = pd.read_excel(input_file, sheet_name='정리표', header=1)
    print(f"   ✓ 정리표 데이터 읽기: {len(df)}개 행")

    # 2. 데이터 변환
    df['아주/날짜'] = df.apply(
        lambda row: format_quantity_date(row['아주'], row['소비기한']),
        axis=1
    )
    df['잔량/날짜'] = df.apply(
        lambda row: format_quantity_date(row['잔량'], row['소비기한']),
        axis=1
    )

    # 같은 제품코드 + 제품명 + 소비기한 그룹화
    grouped = df.groupby(['제품코드', 'Brand Name', '소비기한']).agg({
        '아주/날짜': lambda x: ' '.join([v for v in x if v != '']),
        '잔량/날짜': lambda x: ' '.join([v for v in x if v != ''])
    }).reset_index()

    # 결과 DataFrame
    result = pd.DataFrame()
    result['제품코드'] = grouped['제품코드']
    result['제품명'] = grouped['Brand Name']
    result['아주/날짜'] = grouped['아주/날짜']
    result['일반/날짜'] = ""
    result['잔량/날짜'] = grouped['잔량/날짜']

    result = result[
        (result['아주/날짜'] != "") | (result['잔량/날짜'] != "")
    ].copy()

    print(f"   ✓ 데이터 변환 완료: {len(result)}개 제품")

    # 3. 엑셀 파일 열기
    wb = load_workbook(input_file)

    # 재고표_출력 시트가 있으면 삭제
    if '재고표_출력' in wb.sheetnames:
        del wb['재고표_출력']
        print("   ✓ 기존 재고표_출력 시트 삭제")

    # 새 시트 생성
    ws = wb.create_sheet('재고표_출력')
    print("   ✓ 재고표_출력 시트 생성")

    # 4. 양식 파일에서 열 너비 복사
    wb_template = load_workbook(template_file)
    ws_template = wb_template['콜라']

    # 열 너비 복사
    ws.column_dimensions['A'].width = ws_template.column_dimensions['A'].width  # 제품코드
    ws.column_dimensions['B'].width = ws_template.column_dimensions['B'].width  # 제품명
    ws.column_dimensions['C'].width = ws_template.column_dimensions['C'].width  # 아주/날짜
    ws.column_dimensions['D'].width = ws_template.column_dimensions['D'].width  # 일반/날짜
    ws.column_dimensions['E'].width = ws_template.column_dimensions['E'].width  # 잔량/날짜

    print("   ✓ 열 너비 복사 완료")

    # 5. 헤더 작성
    headers = ['제품코드', '제품명', '아주/날짜', '일반/날짜', '잔량/날짜']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name='맑은 고딕', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 행 높이 설정
    ws.row_dimensions[1].height = 20

    # 6. 데이터 작성
    for idx, row in result.iterrows():
        row_num = idx + 2  # 헤더 다음 행부터

        # 제품코드
        cell = ws.cell(row=row_num, column=1, value=row['제품코드'])
        cell.font = Font(name='맑은 고딕', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 제품명
        cell = ws.cell(row=row_num, column=2, value=row['제품명'])
        cell.font = Font(name='맑은 고딕', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # 아주/날짜 (수량 강조)
        cell = ws.cell(row=row_num, column=3, value=row['아주/날짜'])
        cell.font = Font(name='맑은 고딕', size=12, bold=True)  # 크고 굵게
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 일반/날짜
        cell = ws.cell(row=row_num, column=4, value=row['일반/날짜'])
        cell.font = Font(name='맑은 고딕', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 잔량/날짜 (수량 강조)
        cell = ws.cell(row=row_num, column=5, value=row['잔량/날짜'])
        cell.font = Font(name='맑은 고딕', size=12, bold=True)  # 크고 굵게
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 행 높이
        ws.row_dimensions[row_num].height = 18

    print(f"   ✓ 데이터 작성 완료: {len(result)}개 행")

    # 7. 저장
    wb.save(input_file)
    print(f"   ✓ 파일 저장 완료: {input_file}")

    print("\n" + "="*70)
    print("✅ 재고표 시트 생성 완료!")
    print("="*70)
    print(f"파일: {input_file}")
    print(f"시트: 재고표_출력")
    print(f"총 {len(result)}개 제품")
    print("\n엑셀 파일을 열어서 '재고표_출력' 시트를 확인하세요.")


if __name__ == "__main__":
    input_file = "본두리편의점 재고실사.xlsx"
    template_file = "재고실사양식표.xlsx"

    try:
        create_inventory_sheet_in_excel(input_file, template_file)
    except FileNotFoundError as e:
        print(f"❌ 오류: 파일을 찾을 수 없습니다.")
        print(f"   {e}")
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
